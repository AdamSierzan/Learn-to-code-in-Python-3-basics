
#to work with excel sheets in VSC and python, we need to load one of the modules for xlsx. file
import openpyxl 

wb_obj = openpyxl.load_workbook('original.xlsx')#now, we load the file with the excel sheet


#we can now perform different actions, 
#to check the active object
sheet_obj = wb_obj.active
print(sheet_obj)

#we want to get the object 
cell_obj = sheet_obj.cell(row=1, column = 1)
print(cell_obj)

#for value 
print(cell_obj.value)

#to get the total number of columns and rows
print(sheet_obj.max_row)
print(sheet_obj.max_column)

#to print all colmumn names
max_column = sheet_obj.max_column

for i in range(1, max_column+1):
    cell_obj = sheet_obj.cell(row=1, column = i)
    print(cell_obj.value)

max_row  = sheet_obj.max_row

for i in range(1, max_row+1):
    cell_obj1 = sheet_obj.cell(row=i, column = 1)
    print(cell_obj1.value)
    print(sales)