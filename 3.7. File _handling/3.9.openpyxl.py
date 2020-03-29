import openpyxl
wb = openpyxl.load_workbook('original.xlsx')
print(type(wb))
print(wb.get_sheet_names())

#this fuction is used to get access to the particular sheet
sheet2 = wb.get_sheet_by_name('customers')
print(sheet2)
print(wb.active)

#to access data from sheet cells we refer b sheet and
#then the sheet adress
print(sheet2['A2'].value)

#or

e=sheet2['B2']
print(e.value)
print(e.row) #to get the row
print(e.column) #to get the column

#to get the data from cells with help of rows and columns
print(sheet2.cell(row=2, column=3).value)
#to print the whole column we use "for, in"
for x in range (1,5):
    print(x, sheet2.cell(row = x, column=1).value)
for y in range (1,6):
    print(y, sheet2.cell(row=1, column=y).value)
for z in range (1,5):
    print(z,sheet2.cell(row=z, column=2).value)
for a in len(column):
    print(a,sheet2.cell(row=a, column=3.value))